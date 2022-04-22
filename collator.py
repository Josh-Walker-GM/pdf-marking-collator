
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Series, Reference

import glob
import os
import collections
import statistics
import argparse
import logging

class MarkComment():
    raw_annotation = None
    author: str = None
    question_id: str = None
    mark: float = None
    
    def __init__(self, raw_annotation):
        self.raw_annotation = raw_annotation
        self.author = raw_annotation.info["title"].strip()
        self.question_id = raw_annotation.info["content"].strip().split(" ")[1]
        self.mark = float(raw_annotation.info["content"].strip().split(" ")[2])

    def __str__(self):
        return "Question: {}, Mark: {}, Author: {}".format(self.question_id, self.mark, self.author)

class FeedbackComment():
    raw_annotation = None
    author: str = None
    text: str = None
    page: int = None
    flags = None
    rect = None
    type = None

    def __init__(self, raw_annotation):
        self.raw_annotation = raw_annotation
        self.author = raw_annotation.info["title"].strip()
        self.text = raw_annotation.info["content"].strip()
        self.page = raw_annotation.parent.number
        self.rect = [raw_annotation.rect[0], raw_annotation.rect[1], raw_annotation.rect[2], raw_annotation.rect[3]]
        self.flags = raw_annotation.flags
        self.type = raw_annotation.type[1]

def get_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_dir", metavar="input-dir", type=str, help="directory of pdf collection")
    parser.add_argument("input_file", metavar="input-file", type=str, help="name of base pdf")
    parser.add_argument("--output-file", type=str, help="name of output pdf", default="output.pdf")
    parser.add_argument("--comment-prefix-flag", type=str, help="comment prefix which flags marks", default="!#")
    parser.add_argument("--alias-authors", type=bool, help="replace author names with alias", default=True, action=argparse.BooleanOptionalAction)
    parser.add_argument("--generate-spreadsheet", type=bool, help="generate spreadsheet of extracted marks", default=False, action=argparse.BooleanOptionalAction)
    # parser.add_argument("--use_spreadsheet", type=bool, help="use spreadsheet of marks inplace of pdf markings")
    return parser.parse_args()

def generate_spreadsheet(args, authors: list[str], aliases: list[str], all_marks: list[list[MarkComment]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Marks"
    
    ws.cell(column=2, row=2, value="Alias")
    ws.cell(column=2, row=3, value="Authors")
    for col in range(len(authors)):
        ws.cell(column=col+3, row=3, value="{0}".format(authors[col]))
        if args.alias_authors:
            ws.cell(column=col+3, row=2, value="{0}".format(aliases[authors[col]]))

    question_ids: list[str] = []
    for document_marks in all_marks:
        for mark in document_marks:
            if mark.question_id not in question_ids:
                question_ids.append(mark.question_id)
    question_ids.sort()

    for i in range(len(question_ids)):
        ws.cell(column=2, row=4+i, value="Q: {}".format(question_ids[i]))
    ws.cell(column=2, row=4+len(question_ids)+1, value="Total")
    
    for i in range(len(authors)):
        ws.cell(column=i+3, row=4+len(question_ids)+1).value = "=SUM({}{}:{}{})".format(get_column_letter(3+i), 4, get_column_letter(3+i), 3+len(question_ids))
    ws.cell(column=4+len(authors), row=4+len(question_ids)+1).value = "=SUM({}{}:{}{})".format(get_column_letter(4+len(authors)), 4, get_column_letter(4+len(authors)), 3+len(question_ids))

    for document_marks in all_marks:
        for mark in document_marks:#
            row_index = question_ids.index(mark.question_id)
            column_index = authors.index(mark.author)
            ws.cell(column=3+column_index, row=4+row_index).value = mark.mark

    ws.cell(column=3+len(authors)+1, row=3, value="Average")
    for i in range(len(question_ids)): 
        ws.cell(column=3+len(authors)+1, row=4+i, value="=AVERAGE({}{}:{}{})".format("C", 4+i, get_column_letter(2+len(authors)), 4+i))

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = "Mark Given"
    chart.x_axis.title = "Question ID"

    data = Reference(ws, min_col=3, min_row=3, max_row=len(question_ids)+3, max_col=len(authors)+2)
    cats = Reference(ws, min_col=2, min_row=4, max_row=len(question_ids)+3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "{}{}".format(get_column_letter(len(authors)+4), 2))

    wb.save(filename = os.path.join(os.getcwd(), args.input_dir, "extracted_marks.xlsx"))

def main():

    # set logging format
    logging.basicConfig(format='%(asctime)s: %(message)s', datefmt='%d-%b-%y %H:%M:%S', level=logging.INFO)

    # extract agruments using argparse standard lib
    args = get_arguments()

    # validate input directory
    if not os.path.exists(os.path.join(os.getcwd(), args.input_dir)):
        logging.error("Input directory \"{}\" does not exist!".format(os.path.join(os.getcwd(), args.input_dir)))
        exit(-1)

    # validate input file
    if not os.path.exists(os.path.join(os.getcwd(), args.input_dir, args.input_file)):
        logging.error("Input file \"{}\" does not exist!".format(os.path.join(os.getcwd(), args.input_dir, args.input_file)))
        exit(-1)

    logging.info("Collating all pdf's in \"{}\" using \"{}\" as base".format(os.path.join(os.getcwd(), args.input_dir), os.path.join(os.getcwd(), args.input_dir, args.input_file)))

    # get list of pdf files in collection, remove base and output files
    pdf_collection = glob.glob(os.path.join(os.getcwd(), args.input_dir, "*.pdf"))
    pdf_collection.remove(os.path.join(os.getcwd(), args.input_dir, args.input_file))
    if os.path.join(os.getcwd(), args.input_dir, args.output_file) in pdf_collection:
        pdf_collection.remove(os.path.join(os.getcwd(), args.input_dir, args.output_file))

    all_marks: list[list[MarkComment]] = []
    all_comments: list[list[FeedbackComment]] = []

    for pdf in pdf_collection:
        logging.debug("Reading \"{}\"".format(pdf))
        document = fitz.open(pdf)
        document_marks: list[MarkComment] = []
        document_comments: list[FeedbackComment] = []
        for page in document:
            for annotation in page.annots():
                if annotation.info["content"].strip().startswith(args.comment_prefix_flag):
                    document_marks.append(MarkComment(annotation))
                else:
                    document_comments.append(FeedbackComment(annotation))
        document.close()
        all_marks.append(document_marks)
        all_comments.append(document_comments)

    authors: list[str] = []
    total_comments = 0
    for document_marks in all_marks:
        for mark in document_marks:
            if mark.author not in authors:
                authors.append(mark.author)
    for document_comments in all_comments:
        total_comments += len(document_comments)
        for comment in document_comments:
            if comment.author not in authors:
                authors.append(comment.author)

    logging.info("Extracted {} comments from {} authors from {} files.".format(total_comments, len(authors), len(pdf_collection)))

    if args.alias_authors:
        logging.info("Replacing author names.")
        aliases: dict[str, str] = {}
        for i in range(len(authors)):
            aliases[authors[i]] = "Marker #{}".format(i+1)

    averaged_marks: dict[str, float] = {}
    extracted_marks: dict[str, list[float]] = {}
    total_averaged_mark = 0.0
    for document_marks in all_marks:
        for mark in document_marks:
            if mark.question_id not in extracted_marks:
                extracted_marks[mark.question_id] = []
            extracted_marks[mark.question_id].append(mark.mark)
    for question in extracted_marks:
        averaged_marks[question] = statistics.mean(extracted_marks[question])
        total_averaged_mark += averaged_marks[question]

    averaged_marks = collections.OrderedDict(sorted(averaged_marks.items()))

    document = fitz.open(os.path.join(os.getcwd(), args.input_dir, args.input_file))
    for document_comments in all_comments:
        for comment in document_comments:
            page = document[comment.page]
            if comment.type == "Text" or comment.type == "FreeText":
                annotation = page.add_text_annot([comment.rect[0], comment.rect[1]], comment.text, "Comment")
            elif comment.type == "Highlight":
                annotation = page.add_highlight_annot(comment.rect)
            elif comment.type == "StrikeOut":
                annotation = page.add_strikeout_annot(comment.rect)
            else:
                logging.warning("Annotation of type {} is not supported.".format(comment.type))
                continue
            if args.alias_authors:
                annotation.set_info(content=comment.text, title=aliases[comment.author])
            else:
                annotation.set_info(content=comment.text, title=comment.author)
            annotation.set_flags(comment.flags)
            annotation.update()

    page = document[0]
    total_mark_annotation = page.add_text_annot([25.0, 25.0], "Overall mark: {:.2f}".format(total_averaged_mark))
    total_mark_annotation.set_colors({"stroke": (1.0, 0.0, 0.0), "fill": None})
    total_mark_annotation.set_info(title="Markers")
    total_mark_annotation.update()

    index = 0
    for key, value in averaged_marks.items():
        mark_annotation = page.add_text_annot([25.0, 70.0 + 20.0*index], "Question {}: {:.2f}".format(key, value))
        mark_annotation.set_info(title="Markers")
        mark_annotation.set_colors({"stroke": (1.0, 0.0, 0.0), "fill": None})
        mark_annotation.update()
        index += 1

    document.save(os.path.join(os.getcwd(), args.input_dir, args.output_file))
    document.close()

    if args.generate_spreadsheet:
        logging.info("Generating spreadsheet of extracted marks.")
        generate_spreadsheet(args, authors, aliases, all_marks)


if __name__ == "__main__":
    main()
