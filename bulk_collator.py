
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

import glob
import os
import subprocess
import collections
import statistics
import argparse
import logging


def get_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("directory_list", metavar="directory_list", nargs='+', help="list of directories to collate together")
    parser.add_argument("--generate-individual-spreadsheet", type=bool, default=False, action=argparse.BooleanOptionalAction, help="generate marks spreadsheet for individual collations")
    parser.add_argument("--use-individual-spreadsheet", type=bool, default=False, action=argparse.BooleanOptionalAction, help="use marks spreadsheet to override pdf marks for individual collations")
    parser.add_argument("--generate-combined-spreadsheet", type=bool, default=False, action=argparse.BooleanOptionalAction, help="generate spreadsheet of all collated marks")
    parser.add_argument("--use-combined-spreadsheet", type=bool, default=False, action=argparse.BooleanOptionalAction, help="use combined spreadsheet to override marks from collated pdfs")
    return parser.parse_args()


def generate_combined_spreadsheet(args):
    combined_wb = Workbook()
    combined_ws = combined_wb.active

    row_offset = 2
    column_offset = 2

    directory_index = 0
    for directory in args.directory_list:
        individual_wb = load_workbook(os.path.join(directory, "extracted_marks.xlsx"))
        individual_ws = individual_wb.active

        marker_count = 0
        while True:
            if individual_ws.cell(2, marker_count + 3).value is None:
                break
            marker_count += 1

        question_count = 0
        while True:
            if individual_ws.cell(question_count + 4, 2).value is None:
                break
            question_count += 1

        combined_ws.cell(row_offset, column_offset).value = directory
        for row in range(question_count + 2):
            for column in range(marker_count + 1):
                combined_ws.cell(row_offset + row + 1, column_offset + column).value = individual_ws.cell(row + 2, column + 2).value

        combined_ws.cell(row_offset + question_count + 4, column_offset).value = "Total"
        for column in range(marker_count):
            combined_ws.cell(row_offset + question_count + 4, column_offset + column + 1).value = "=SUM({}{}:{}{})".format(get_column_letter(column_offset + column + 1), row_offset + 3, get_column_letter(column_offset + column + 1), row_offset + 2 + question_count)
        combined_ws.cell(row_offset + question_count + 4, column_offset + marker_count + 2).value = "=SUM({}{}:{}{})".format(get_column_letter(column_offset + marker_count + 2), row_offset + 3, get_column_letter(column_offset + marker_count + 2), row_offset + 2 + question_count)
        combined_ws.cell(row_offset + 2, column_offset + marker_count + 2).value = "Average"
        for row in range(question_count):
            combined_ws.cell(row_offset + row + 3, column_offset + marker_count + 2).value = "=AVERAGE({}{}:{}{})".format(get_column_letter(column_offset + 1), row_offset + row + 3, get_column_letter(column_offset + marker_count), row_offset + row + 3)

        # create bar chart for marking data visualisation
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.y_axis.title = "Mark Given"
        chart.x_axis.title = "Question ID"
        data = Reference(combined_ws, min_col=column_offset+1, min_row=row_offset+2, max_row=row_offset+question_count+2, max_col=column_offset+marker_count)
        cats = Reference(combined_ws, min_col=column_offset, min_row=row_offset+3, max_row=row_offset+question_count+2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 0.55 * (question_count + 5)
        chart.width = 3 * (question_count)
        combined_ws.add_chart(chart, "{}{}".format(get_column_letter(column_offset + marker_count + 4), row_offset))

        # for i in range(len(authors)):
        #     ws.cell(column=i+3, row=4+len(question_ids)+1).value = "=SUM({}{}:{}{})".format(get_column_letter(3+i), 4, get_column_letter(3+i), 3+len(question_ids))
        # ws.cell(column=4+len(authors), row=4+len(question_ids)+1).value = "=SUM({}{}:{}{})".format(get_column_letter(4+len(authors)), 4, get_column_letter(4+len(authors)), 3+len(question_ids))

        directory_index += 1
        row_offset += question_count + 6

    combined_wb.save(filename=os.path.join(os.getcwd(), "combined_extacted_marks.xlsx"))

# def read_spreadsheet(args, authors, question_ids):
#     wb = load_workbook(os.path.join(os.getcwd(), args.input_dir, "extracted_marks.xlsx"))
#     ws = wb.active

#     # read grid of marks
#     overriding_marks: list[list[MarkComment]] = []
#     for col in range(len(authors)):
#         marks: list[MarkComment] = []
#         for row in range(len(question_ids)):
#             mark_comment = MarkComment(None)
#             mark_comment.author = authors[col]
#             mark_comment.question_id = question_ids[row]
#             mark_comment.mark = ws.cell(row+4, col+3).value
#             marks.append(mark_comment)
#         overriding_marks.append(marks)

#     return overriding_marks


def main():

    # set logging format
    logging.basicConfig(format='%(asctime)s: %(message)s', datefmt='%d-%b-%y %H:%M:%S', level=logging.INFO)

    # extract agruments using argparse standard lib
    args = get_arguments()

    # validate against usage of override with and generate spreadsheet features together
    if args.generate_combined_spreadsheet and args.use_combined_spreadsheet:
        logging.error("Cannot use overriding spreadsheet and generate spreadsheet features at the same time!")
        exit(-1)

    # validate against usage of override with and generate individual spreadsheet flags together
    if args.generate_individual_spreadsheet and args.use_individual_spreadsheet:
        logging.error("Cannot use both use and generate individual spreadsheet flags together!")
        exit(-1)

    # validate all collation directories are unique
    if len(args.directory_list) != len(set(args.directory_list)):
        logging.error("Collation directory list cannot contain duplicates!")
        exit(-1)

    # validate all collation directories exist
    for directory in args.directory_list:
        if not os.path.exists(os.path.join(os.getcwd(), directory)):
            logging.error("Collation directory \"{}\" does not exist!".format(os.path.join(os.getcwd(), directory)))
            exit(-1)

    # Dev note: Calling of commands formed from user input is dangerous - should check/sanitise this...
    # for directory in args.directory_list:
    #     logging.info("Collating {}.".format(directory))
    #     command_string = "python collator.py {} {}.pdf".format(directory, os.path.basename(os.path.normpath(directory)))

    #     if args.generate_individual_spreadsheet:
    #         command_string = "{} {}".format(command_string, "--generate-spreadsheet")

    #     if args.use_individual_spreadsheet:
    #         command_string = "{} {}".format(command_string, "--use-spreadsheet")

    #     return_code = subprocess.call(command_string, shell=True)
    #     if return_code != 0:
    #         logging.error("Collation of \"{}\" failed!".format(directory))
    #         exit(-1)

    if args.generate_combined_spreadsheet:
        logging.info("Generating combined spreadsheet.")

        # validate individual projects have extracted marks spreadsheets
        for directory in args.directory_list:
            if not os.path.exists(os.path.join(directory, "extracted_marks.xlsx")):
                logging.error("Extracted marks spreadsheet does not exist for \"{}\"!".format(directory))
                exit(-1)

        generate_combined_spreadsheet(args)


if __name__ == "__main__":
    main()
